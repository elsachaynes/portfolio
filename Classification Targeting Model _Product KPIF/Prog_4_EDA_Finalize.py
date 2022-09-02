# -*- coding: utf-8 -*-
"""
------------------------------------------------------------------------------
Program Name: Prog_4_EDA_Finalize.py
Date Created: 8/21/2022
Created By: Elsa Haynes (elsa.c.haynes@kp.org)

Outputs documentaiton on variables chosen for model build.

Creates one tab for side-by-side comparison. Creates one tab for each dataset.

------------------------------------------------------------------------------
"""

# %% Initialize Libraries and Paths

import pandas as pd
import os.path
from openpyxl import load_workbook

inputpath = (r'C:\Users\c156934')

outputpath = (r'\\cs.msds.kp.org\scal\REGS\share15\KPIF_ANALYTICS\Elsa'
              r'\KPIF_ Targeting Model 2023 OE')

LOB = 'KPIF'

# %% Import CSV files


def import_csv(csv_name):
    global inputpath
    df = pd.read_csv(inputpath + csv_name + '.csv',
                     dtype={'AGLTY_INDIV_ID': object})
    return(df)


FM_Off = import_csv(r'\Final Model Set FM Off HIX')  # 5,608,463 rows
FM_On = import_csv(r'\Final Model Set FM On HIX')  # 5,608,463 rows
RW_Off = import_csv(r'\Final Model Set RW Off HIX')  # 1,609,869 rows
RW_On = import_csv(r'\Final Model Set RW On HIX')  # 1,609,869 rows


def export_excel(df, file_path, file_sheet_name):
    if os.path.exists(file_path):
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        ws = writer.book[file_sheet_name]
        for row in ws['A1:E200']:
            for cell in row:
                cell.value = 0
        df.to_excel(writer, sheet_name=file_sheet_name, index=False)
        writer.save()
        writer.close()
        # if new
    else:
        df.to_excel(file_path, sheet_name=file_sheet_name, index=False)


# %% Output variable list

ListDF = [FM_Off, FM_On, RW_Off, RW_On]  # manual input required
ColNames = ['FM_Off', 'FM_On', 'RW_Off', 'RW_On']  # manual input required

FinalVars = pd.DataFrame()
output = []
for (df, name) in zip(ListDF, ColNames):
    ListCols = pd.DataFrame(list(df.columns)).rename(columns={0: 'Variable'})
    FinalVars = pd.concat([FinalVars, ListCols])\
        .drop_duplicates(keep='first')
    ListCols[name] = 'Y'
    output.append(ListCols)
    # print(name + 'added ' + {len(ListColsAll)} + ' variables.')

del df
del name
del ListCols

FinalVars = FinalVars.rename(columns={0: 'Variable'})

for (i, name) in zip(output, ColNames):
    FinalVars = pd.merge(FinalVars, i, on='Variable', how='left')

FinalVars = FinalVars.sort_values(by='Variable')

del i
del name
del output
del ListDF
del ColNames

# Export
excel_path = outputpath + '\Final_Tabulations.xlsx'
export_excel(FinalVars, excel_path, 'Variable List')

# %% Output descriptive statistics



